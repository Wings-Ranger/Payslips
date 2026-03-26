import json
import re
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from PyPDF2 import PdfReader
from dateutil import parser as date_parser
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class PayslipRecord:
    file_name: str
    employee: Optional[str]
    pay_date: Optional[str]
    pay_period: Optional[str]
    week_start: Optional[str]
    # Ordinary hours
    ordinary_hours: Optional[float] = None
    ordinary_rate: Optional[float] = None
    ordinary_pay_this: Optional[float] = None
    ordinary_pay_ytd: Optional[float] = None
    # Weekend hours
    weekend_hours: Optional[float] = None
    weekend_rate: Optional[float] = None
    weekend_pay_this: Optional[float] = None
    weekend_pay_ytd: Optional[float] = None
    # Public holiday hours
    public_holiday_hours: Optional[float] = None
    public_holiday_rate: Optional[float] = None
    public_holiday_pay_this: Optional[float] = None
    public_holiday_pay_ytd: Optional[float] = None
    # Totals
    gross_this_pay: Optional[float] = None
    gross_ytd: Optional[float] = None
    tax_this_pay: Optional[float] = None
    tax_ytd: Optional[float] = None
    payg_this_pay: Optional[float] = None
    payg_ytd: Optional[float] = None
    net_this_pay: Optional[float] = None
    net_ytd: Optional[float] = None
    total_hours_this_pay: Optional[float] = None
    notes: str = ""


def load_config(project_root: Path) -> dict:
    config_path = project_root / "src" / "config.json"
    if not config_path.exists():
        raise FileNotFoundError(f"Missing config file: {config_path}")
    with config_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def extract_currency_values(line: str) -> list[float]:
    """Extract all numeric values from a text line as floats."""
    return [float(x) for x in re.findall(r"[\d.]+", line)]


def read_text_from_file(file_path: Path) -> str:
    if file_path.suffix.lower() == ".pdf":
        reader = PdfReader(str(file_path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)

    if file_path.suffix.lower() == ".txt":
        return file_path.read_text(encoding="utf-8", errors="ignore")

    return ""


def get_week_start(dt: datetime, start_day: str = "monday") -> datetime:
    weekdays = {
        "monday": 0,
        "tuesday": 1,
        "wednesday": 2,
        "thursday": 3,
        "friday": 4,
        "saturday": 5,
        "sunday": 6,
    }
    target = weekdays.get(start_day.lower(), 0)
    offset = (dt.weekday() - target) % 7
    return dt - timedelta(days=offset)


def parse_payslip(file_path: Path, text: str, config: dict) -> PayslipRecord:
    # Check if PDF is scanned (no extractable text)
    if len(text.strip()) < 50:
        return PayslipRecord(
            file_name=file_path.name,
            employee=None,
            pay_date=None,
            pay_period=None,
            week_start=None,
            notes="SKIPPED: Scanned PDF, needs OCR.",
        )
    
    # Extract employee name - first non-empty line
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    employee_raw = lines[0] if lines else None

    # Normalize text for regex (join lines, preserve structure)
    text_normalized = text.replace("\n", " ")
    
    # Extract pay period and payment date
    pay_date = None
    pay_period = None
    m = re.search(r"pay period:\s*(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})", text_normalized, re.IGNORECASE)
    if m:
        pay_period = f"{m.group(1)} - {m.group(2)}"
    
    m = re.search(r"payment date:\s*(\d{1,2}/\d{1,2}/\d{4})", text_normalized, re.IGNORECASE)
    if m:
        pay_date_str = m.group(1)
        try:
            pay_date_obj = date_parser.parse(pay_date_str, dayfirst=True)
            pay_date = pay_date_obj.date().isoformat()
        except Exception:
            pass

    # Initialize field values
    ordinary_hours = None
    ordinary_rate = None
    ordinary_pay_this = None
    ordinary_pay_ytd = None
    weekend_hours = None
    weekend_rate = None
    weekend_pay_this = None
    weekend_pay_ytd = None
    public_holiday_hours = None
    public_holiday_rate = None
    public_holiday_pay_this = None
    public_holiday_pay_ytd = None
    gross_this_pay = None
    gross_ytd = None
    payg_this_pay = None
    payg_ytd = None
    tax_this_pay = None
    tax_ytd = None
    net_this_pay = None
    net_ytd = None

    # Parse salary & wages section line by line
    in_salary_section = False
    for line in text.splitlines():
        line_lower = line.lower()
        
        if "salary & wages" in line_lower:
            in_salary_section = True
            continue
        if in_salary_section and ("tax" in line_lower or line_lower.startswith("tax")):
            in_salary_section = False
        
        if in_salary_section:
            # Extract ordinary hours line: "Ordinary Hours 7.5000 $16.7100 $125.32 $7,337.90"
            if "ordinary hours" in line_lower:
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 4:
                    try:
                        ordinary_hours = float(parts[0])
                        ordinary_rate = float(parts[1])
                        ordinary_pay_this = float(parts[2])
                        ordinary_pay_ytd = float(parts[3])
                    except (ValueError, IndexError):
                        pass
            
            # Extract weekend hours line: "Weekends Sat/Sun 5.0000 $21.8000 $109.00 $4,352.33"
            elif "weekends" in line_lower and ("sat" in line_lower or "sun" in line_lower):
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 4:
                    try:
                        weekend_hours = float(parts[0])
                        weekend_rate = float(parts[1])
                        weekend_pay_this = float(parts[2])
                        weekend_pay_ytd = float(parts[3])
                    except (ValueError, IndexError):
                        pass
            
            # Extract public holiday line
            elif "public" in line_lower and "holiday" in line_lower:
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 4:
                    try:
                        public_holiday_hours = float(parts[0])
                        public_holiday_rate = float(parts[1])
                        public_holiday_pay_this = float(parts[2])
                        public_holiday_pay_ytd = float(parts[3])
                    except (ValueError, IndexError):
                        pass
            
            # Extract TOTAL line for gross: "TOTAL $234.32 $12,235.23"
            elif line_lower.strip().startswith("total"):
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 2:
                    try:
                        gross_this_pay = float(parts[0])
                        gross_ytd = float(parts[1])
                    except (ValueError, IndexError):
                        pass

    # Parse TAX section
    in_tax_section = False
    for line in text.splitlines():
        line_lower = line.lower()
        
        if line_lower.startswith("tax"):
            in_tax_section = True
            continue
        if in_tax_section and "payment details" in line_lower:
            in_tax_section = False
        
        if in_tax_section:
            # PAYG line: "PAYG $0.00 $264.00"
            if "payg" in line_lower:
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 2:
                    try:
                        payg_this_pay = float(parts[0])
                        payg_ytd = float(parts[1])
                    except (ValueError, IndexError):
                        pass
            
            # TAX or other tax types
            elif "tax" in line_lower and "payg" not in line_lower and line_lower.strip() != "tax":
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 2:
                    try:
                        tax_this_pay = float(parts[0])
                        tax_ytd = float(parts[1])
                    except (ValueError, IndexError):
                        pass
            
            # TOTAL under tax: "TOTAL $0.00 $264.00"
            elif line_lower.strip().startswith("total"):
                parts = re.findall(r"[\d.]+", line)
                if len(parts) >= 2 and tax_this_pay is None:  # Only set if not already set
                    try:
                        tax_this_pay = float(parts[0])
                        tax_ytd = float(parts[1])
                    except (ValueError, IndexError):
                        pass

    # Extract Net Pay: "Net Pay: $234.32"
    m = re.search(r"net pay:\s*\$?([\d.]+)", text_normalized, re.IGNORECASE)
    if m:
        try:
            net_this_pay = float(m.group(1))
        except ValueError:
            pass

    # Total hours this pay = ordinary + weekend + public holiday
    total_hours_this_pay = 0.0
    if ordinary_hours:
        total_hours_this_pay += ordinary_hours
    if weekend_hours:
        total_hours_this_pay += weekend_hours
    if public_holiday_hours:
        total_hours_this_pay += public_holiday_hours
    total_hours_this_pay = total_hours_this_pay if total_hours_this_pay > 0 else None

    # Calculate week_start from pay_date
    week_start = None
    if pay_date:
        try:
            pay_date_obj = datetime.fromisoformat(pay_date)
            week_start = get_week_start(pay_date_obj, config.get("week_start_day", "monday")).date().isoformat()
        except Exception:
            pass

    notes = []
    if pay_date is None:
        notes.append("Could not determine pay date")

    return PayslipRecord(
        file_name=file_path.name,
        employee=employee_raw,
        pay_date=pay_date,
        pay_period=pay_period,
        week_start=week_start,
        ordinary_hours=ordinary_hours,
        ordinary_rate=ordinary_rate,
        ordinary_pay_this=ordinary_pay_this,
        ordinary_pay_ytd=ordinary_pay_ytd,
        weekend_hours=weekend_hours,
        weekend_rate=weekend_rate,
        weekend_pay_this=weekend_pay_this,
        weekend_pay_ytd=weekend_pay_ytd,
        public_holiday_hours=public_holiday_hours,
        public_holiday_rate=public_holiday_rate,
        public_holiday_pay_this=public_holiday_pay_this,
        public_holiday_pay_ytd=public_holiday_pay_ytd,
        gross_this_pay=gross_this_pay,
        gross_ytd=gross_ytd,
        tax_this_pay=tax_this_pay,
        tax_ytd=tax_ytd,
        payg_this_pay=payg_this_pay,
        payg_ytd=payg_ytd,
        net_this_pay=net_this_pay,
        net_ytd=net_ytd,
        total_hours_this_pay=total_hours_this_pay,
        notes="; ".join(notes),
    )


def find_missing_weeks(df: pd.DataFrame) -> list[str]:
    if df.empty or "week_start" not in df.columns:
        return []

    valid = df["week_start"].dropna().unique().tolist()
    if not valid:
        return []

    weeks = sorted(datetime.fromisoformat(x).date() for x in valid)
    start = weeks[0]
    end = weeks[-1]

    observed = set(weeks)
    missing = []

    current = start
    while current <= end:
        if current not in observed:
            missing.append(current.isoformat())
        current += timedelta(days=7)

    return missing


def format_excel_output(xlsx_path: Path) -> None:
    """Apply formatting to Excel output for readability."""
    from openpyxl import load_workbook
    
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    
    # Define colors for different sections
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    ordinary_fill = PatternFill(start_color="E7E6FF", end_color="E7E6FF", fill_type="solid")
    weekend_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    public_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gross_fill = PatternFill(start_color="FDB766", end_color="FDB766", fill_type="solid")
    tax_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    col_widths = {
        'A': 20,  # file_name
        'B': 18,  # employee
        'C': 12,  # pay_date
        'D': 18,  # pay_period
        'E': 12,  # week_start
        'F': 15,  # ordinary_hours
        'G': 13,  # ordinary_rate
        'H': 16,  # ordinary_pay_this
        'I': 15,  # ordinary_pay_ytd
        'J': 14,  # weekend_hours
        'K': 13,  # weekend_rate
        'L': 16,  # weekend_pay_this
        'M': 15,  # weekend_pay_ytd
        'N': 16,  # public_holiday_hours
        'O': 15,  # public_holiday_rate
        'P': 18,  # public_holiday_pay_this
        'Q': 17,  # public_holiday_pay_ytd
        'R': 15,  # gross_this_pay
        'S': 12,  # gross_ytd
        'T': 14,  # tax_this_pay
        'U': 11,  # tax_ytd
        'V': 15,  # payg_this_pay
        'W': 12,  # payg_ytd
        'X': 14,  # net_this_pay
        'Y': 11,  # net_ytd
        'Z': 18,  # total_hours_this_pay
        'AA': 16, # notes
    }

    # Apply column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set notes column as frozen on the right with reduced width
    ws.column_dimensions['AA'].width = 15

    # Map columns to fill colors
    color_map = {
        'F': ordinary_fill, 'G': ordinary_fill, 'H': ordinary_fill, 'I': ordinary_fill,  # Ordinary
        'J': weekend_fill, 'K': weekend_fill, 'L': weekend_fill, 'M': weekend_fill,      # Weekend
        'N': public_fill, 'O': public_fill, 'P': public_fill, 'Q': public_fill,          # Public holiday
        'R': gross_fill, 'S': gross_fill,                                                # Gross
        'T': tax_fill, 'U': tax_fill, 'V': tax_fill, 'W': tax_fill,                      # Tax
    }
    
    # Format header row
    for col_num, cell in enumerate(ws[1], 1):
        col_letter = get_column_letter(col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Format data rows
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        for col_num, cell in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            
            # Apply section colors
            if col_letter in color_map:
                cell.fill = color_map[col_letter]

            # Apply borders
            cell.border = thin_border
            
            # Format currency columns
            if col_letter in ['H', 'I', 'L', 'M', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # Format numeric columns (hours, rates)
            elif col_letter in ['F', 'G', 'J', 'K', 'N', 'O', 'Z']:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # Format date columns
            elif col_letter in ['C', 'E']:
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Notes column: no wrap, left-aligned
            elif col_letter == 'AA':
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # Center-align text columns
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Freeze header row
    ws.freeze_panes = 'A2'

    # Format missing_weeks sheet
    if "missing_weeks" in wb.sheetnames:
        mw = wb["missing_weeks"]
        mw.column_dimensions['A'].width = 16
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        for cell in mw[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row in mw.iter_rows(min_row=2, max_row=mw.max_row):
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    cell.fill = black_fill
                else:
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        mw.freeze_panes = 'A2'

    wb.save(str(xlsx_path))







# Mapping of backend field names to human-readable Excel headers
EXCEL_HEADERS = {
    "file_name": "File Name",
    "employee": "Employee",
    "pay_date": "Pay Date",
    "pay_period": "Pay Period",
    "week_start": "Week Start",
    "ordinary_hours": "Ordinary Hours",
    "ordinary_rate": "Ordinary Rate",
    "ordinary_pay_this": "Ordinary Pay (This)",
    "ordinary_pay_ytd": "Ordinary Pay (YTD)",
    "weekend_hours": "Weekend Hours",
    "weekend_rate": "Weekend Rate",
    "weekend_pay_this": "Weekend Pay (This)",
    "weekend_pay_ytd": "Weekend Pay (YTD)",
    "public_holiday_hours": "Public Holiday Hours",
    "public_holiday_rate": "Public Holiday Rate",
    "public_holiday_pay_this": "Public Holiday Pay (This)",
    "public_holiday_pay_ytd": "Public Holiday Pay (YTD)",
    "gross_this_pay": "Gross Pay (This)",
    "gross_ytd": "Gross Pay (YTD)",
    "tax_this_pay": "Tax (This)",
    "tax_ytd": "Tax (YTD)",
    "payg_this_pay": "PAYG (This)",
    "payg_ytd": "PAYG (YTD)",
    "net_this_pay": "Net Pay (This)",
    "net_ytd": "Net Pay (YTD)",
    "total_hours_this_pay": "Total Hours",
    "notes": "Notes",
}


REQUIRED_SCHEMA_FIELDS = [
    "file_name",
    "employee",
    "pay_date",
    "week_start",
    "net_this_pay",
]


def validate_record_schema(record: PayslipRecord) -> list[str]:
    """Return list of missing required fields for a parsed record."""
    data = asdict(record)
    missing: list[str] = []
    for field_name in REQUIRED_SCHEMA_FIELDS:
        value = data.get(field_name)
        if value is None:
            missing.append(field_name)
            continue
        if isinstance(value, str) and not value.strip():
            missing.append(field_name)
    return missing


def append_validation_notes(record: PayslipRecord) -> PayslipRecord:
    """Add schema validation errors into record notes for downstream visibility."""
    missing = validate_record_schema(record)
    if not missing:
        return record

    schema_note = f"SCHEMA_INVALID: missing required fields: {', '.join(missing)}"
    if schema_note in (record.notes or ""):
        return record

    if record.notes:
        record.notes = f"{record.notes}; {schema_note}"
    else:
        record.notes = schema_note
    return record


def rename_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Rename dataframe columns to human-readable headers for Excel output."""
    rename_map = {col: EXCEL_HEADERS.get(col, col) for col in df.columns}
    return df.rename(columns=rename_map)


def add_pay_validation_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add cross-check columns that verify parsed pay figures are internally consistent.

    For each record the function computes:
    - ``ordinary_check``  — hours × rate ≈ pay_this  (within $0.02 rounding tolerance)
    - ``weekend_check``   — same check for weekend pay
    - ``public_holiday_check`` — same check for public holiday pay
    - ``gross_check``     — sum of all three *_pay_this ≈ gross_this_pay
    - ``net_check``       — gross − tax − payg ≈ net_this_pay
    - ``overall_pay_check`` — PASS only when every individual check above is PASS

    Values are ``"PASS"``, ``"FAIL"``, or ``"N/A"`` when the inputs needed for the
    check are absent or non-numeric.
    """
    _TOL = 0.02

    def _numeric(val) -> float | None:
        try:
            v = float(val)
            return v if v == v else None  # reject NaN
        except (TypeError, ValueError):
            return None

    def _hours_rate_check(hours_val, rate_val, pay_val) -> str:
        h, r, p = _numeric(hours_val), _numeric(rate_val), _numeric(pay_val)
        if h is None or r is None or p is None:
            return "N/A"
        return "PASS" if abs(h * r - p) <= _TOL else "FAIL"

    out = df.copy()

    out["ordinary_check"] = [
        _hours_rate_check(r["ordinary_hours"], r["ordinary_rate"], r["ordinary_pay_this"])
        for _, r in df.iterrows()
    ]
    out["weekend_check"] = [
        _hours_rate_check(r["weekend_hours"], r["weekend_rate"], r["weekend_pay_this"])
        for _, r in df.iterrows()
    ]
    out["public_holiday_check"] = [
        _hours_rate_check(r["public_holiday_hours"], r["public_holiday_rate"], r["public_holiday_pay_this"])
        for _, r in df.iterrows()
    ]

    gross_checks = []
    net_checks = []
    for _, r in df.iterrows():
        ord_pay = _numeric(r["ordinary_pay_this"]) or 0.0
        wk_pay = _numeric(r["weekend_pay_this"]) or 0.0
        ph_pay = _numeric(r["public_holiday_pay_this"]) or 0.0
        gross = _numeric(r["gross_this_pay"])
        if gross is None:
            gross_checks.append("N/A")
        else:
            gross_checks.append("PASS" if abs(ord_pay + wk_pay + ph_pay - gross) <= _TOL else "FAIL")

        tax = _numeric(r["tax_this_pay"]) or 0.0
        payg = _numeric(r["payg_this_pay"]) or 0.0
        net = _numeric(r["net_this_pay"])
        if gross is None or net is None:
            net_checks.append("N/A")
        else:
            net_checks.append("PASS" if abs(gross - tax - payg - net) <= _TOL else "FAIL")

    out["gross_check"] = gross_checks
    out["net_check"] = net_checks

    overall = []
    check_cols = ["ordinary_check", "weekend_check", "public_holiday_check", "gross_check", "net_check"]
    for _, r in out.iterrows():
        values = [r[c] for c in check_cols]
        if any(v == "FAIL" for v in values):
            overall.append("FAIL")
        elif all(v == "PASS" for v in values):
            overall.append("PASS")
        else:
            overall.append("N/A")
    out["overall_pay_check"] = overall

    return out


def _write_excel(xlsx_path: Path, df: pd.DataFrame, missing_weeks: list[str]) -> None:
    """Write the main payslips sheet and the missing_weeks sheet, then apply formatting."""
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        rename_for_excel(df).to_excel(writer, index=False, sheet_name="payslips")
        pd.DataFrame({"missing_week_start": missing_weeks}).to_excel(
            writer, index=False, sheet_name="missing_weeks", header=["Week Start"]
        )
    format_excel_output(xlsx_path)


def run() -> None:
    project_root = Path(__file__).resolve().parents[1]
    config = load_config(project_root)

    input_dir = project_root / config.get("input_dir", "input")
    output_dir = project_root / config.get("output_dir", "output")
    output_dir.mkdir(parents=True, exist_ok=True)
    input_dir.mkdir(parents=True, exist_ok=True)

    supported = {ext.lower() for ext in config.get("supported_extensions", [".pdf", ".txt"])}
    files = [f for f in input_dir.iterdir() if f.is_file() and f.suffix.lower() in supported]

    if not files:
        print(f"No payslip files found in: {input_dir}")
        print("Add PDF/TXT payslips to input/ and re-run.")
        return

    records: list[PayslipRecord] = []

    for file_path in sorted(files):
        text = read_text_from_file(file_path)
        record = parse_payslip(file_path, text, config)
        record = append_validation_notes(record)
        records.append(record)

    df = pd.DataFrame([asdict(r) for r in records])
    df = df.sort_values(by=["week_start", "pay_date", "file_name"], na_position="last").reset_index(drop=True)

    # Fill pay fields with N/A when not applicable
    ph_cols = ["public_holiday_hours", "public_holiday_rate", "public_holiday_pay_this", "public_holiday_pay_ytd"]
    wk_cols = ["weekend_hours", "weekend_rate", "weekend_pay_this", "weekend_pay_ytd"]
    ord_cols = ["ordinary_hours", "ordinary_rate", "ordinary_pay_this", "ordinary_pay_ytd"]
    df[ph_cols + wk_cols + ord_cols] = df[ph_cols + wk_cols + ord_cols].fillna("N/A")

    missing_weeks = find_missing_weeks(df)

    xlsx_path = output_dir / config.get("output_filename", "payslips.xlsx")
    csv_path = output_dir / "payslips.csv"

    # Try to write Excel file; if locked, use timestamped backup
    try:
        _write_excel(xlsx_path, df, missing_weeks)
    except PermissionError:
        from datetime import datetime as dt
        backup_name = f"payslips_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        xlsx_path = output_dir / backup_name
        _write_excel(xlsx_path, df, missing_weeks)
        print(f"(Note: Main file was locked, saved as: {backup_name})")

    df.to_csv(csv_path, index=False)

    print(f"Processed {len(df)} payslip file(s)")
    print(f"Spreadsheet: {xlsx_path}")
    print(f"CSV: {csv_path}")

    if missing_weeks:
        print("Missing weekly payslips detected:")
        for w in missing_weeks:
            print(f" - {w}")
    else:
        print("No missing weekly payslips detected in the observed range.")


if __name__ == "__main__":
    run()
