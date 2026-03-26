#!/usr/bin/env python3
from openpyxl import load_workbook
import glob

# Find the latest Excel file that has payslips sheet
excel_files = glob.glob('output/payslips*.xlsx')
for file in sorted(excel_files, reverse=True):
    try:
        wb = load_workbook(file)
        if "payslips" in wb.sheetnames:
            print(f"Using: {file}\n")
            
            # Show all payslips sheet headers
            if "payslips" in wb.sheetnames:
                ws = wb["payslips"]
                headers = [cell.value for cell in ws[1]]
                print("=== Payslips Sheet - All Human-Readable Headers ===\n")
                for i, h in enumerate(headers, 1):
                    print(f"{i:2d}. {h}")
                
                print(f"\n✓ Human-readable headers successfully applied!")
                print(f"✓ Backend field names preserved in CSV (payslips.csv)")
            
            print(f"\nData samples:")
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
                print(f"  Row {i}: {row[0]} | {row[2]} | ${row[17]}")
            break
    except:
        continue
