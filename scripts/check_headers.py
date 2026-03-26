#!/usr/bin/env python3
from openpyxl import load_workbook
import glob

# Find the latest Excel file
excel_files = glob.glob('output/payslips*.xlsx')
if not excel_files:
    print("No Excel files found")
    exit(1)

latest_file = max(excel_files)
print(f"Checking: {latest_file}\n")

wb = load_workbook(latest_file)

# Check payslips sheet headers
if "payslips" in wb.sheetnames:
    ws = wb["payslips"]
    headers = [cell.value for cell in ws[1]]
    print("=== Payslips Sheet Headers ===")
    for i, h in enumerate(headers[:10], 1):
        print(f"{i:2d}. {h}")
    print(f"... ({len(headers)} total columns)")
    
print(f"\nAll sheets: {wb.sheetnames}")
