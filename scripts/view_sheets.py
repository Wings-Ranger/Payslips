#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('output/payslips.xlsx')

# Check summary sheet
print("=== SUMMARY SHEET ===")
ws = wb['summary']
for row in ws.iter_rows(values_only=True):
    print(row)

print("\n=== MONTHLY SHEET ===")
ws = wb['monthly']
for row in ws.iter_rows(values_only=True):
    print(row)

print("\n=== WEEKLY SHEET ===")
ws = wb['weekly']
for row in ws.iter_rows(values_only=True):
    print(row)

print("\n=== DEDUCTIONS SHEET ===")
ws = wb['deductions']
for row in ws.iter_rows(values_only=True):
    print(row)

print("\nAll sheets:", wb.sheetnames)
