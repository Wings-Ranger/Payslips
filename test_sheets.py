#!/usr/bin/env python3
from openpyxl import load_workbook
import sys

wb = load_workbook('output/payslips.xlsx')
print('Sheets in workbook:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = len([row for row in ws.iter_rows()])
    cols = ws.max_column
    print(f"\n{sheet_name}: {rows} rows, {cols} columns")
    if rows > 0:
        headers = [cell.value for cell in ws[1]]
        print(f"  Headers: {headers}")
